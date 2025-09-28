"""Excel export engine with automatic chart generation."""

import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .config import settings
from .filename_generator import filename_generator
from .database import format_currency, tiyin_to_sum


class ExcelExporter:
    """Professional Excel export with automatic chart generation."""

    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = Path(output_dir or settings.excel_output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Professional styling colors (Banking theme)
        self.colors = {
            'primary': '1f4e79',      # Dark blue
            'secondary': '5b9bd5',     # Light blue
            'accent': '70ad47',        # Green
            'success': '70ad47',       # Green for positive metrics
            'warning': 'ffc000',       # Orange
            'header': 'f2f2f2',       # Light gray
            'text': '404040'           # Dark gray
        }

        # Multilingual labels
        self.labels = {
            'en': {
                'title': 'Banking Report',
                'data': 'Data',
                'charts': 'Charts',
                'summary': 'Summary',
                'generated': 'Generated on',
                'query': 'Query',
                'records': 'Records'
            },
            'ru': {
                'title': 'Банковский отчет',
                'data': 'Данные',
                'charts': 'Графики',
                'summary': 'Сводка',
                'generated': 'Создан',
                'query': 'Запрос',
                'records': 'Записей'
            },
            'uz': {
                'title': 'Banking hisoboti',
                'data': 'Ma\'lumotlar',
                'charts': 'Diagrammalar',
                'summary': 'Xulosa',
                'generated': 'Yaratilgan',
                'query': 'So\'rov',
                'records': 'Yozuvlar'
            }
        }

    def _detect_language(self, user_query: str) -> str:
        """Simple language detection for Excel labels."""
        query_lower = user_query.lower()

        # Check for Cyrillic characters (Russian)
        if any('\u0400' <= char <= '\u04ff' for char in query_lower):
            return 'ru'

        # Check for common Uzbek terms
        uzbek_terms = ['mijoz', 'hisob', 'operatsiya', 'toshkent', 'samarqand', 'buxoro']
        if any(term in query_lower for term in uzbek_terms):
            return 'uz'

        # Default to English
        return 'en'

    def export_query_results(
        self,
        data: List[Dict],
        query_info: Dict,
        filename: Optional[str] = None
    ) -> str:
        """Export query results to Excel with professional formatting and charts."""

        if not data:
            raise ValueError("No data to export")

        # Create filename if not provided
        if not filename:
            # Use fast filename generator to create descriptive filename
            sql_query = query_info.get('sql_query', '')
            filename = filename_generator.generate_filename(sql_query, data)

        filepath = self.output_dir / filename
        logger.info(f"Exporting data to Excel: {filepath}")

        # Convert to DataFrame for easier manipulation
        df = pd.DataFrame(data)

        # Detect language from original query for proper formatting
        query_language = self._detect_query_language(query_info.get('user_query', ''))

        # Format currency columns with language-specific formatting
        df = self._format_currency_columns(df, query_language)

        # Create workbook and worksheets
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Data"

        # Export main data
        self._write_data_sheet(ws_data, df, query_info, query_language)


        # Add basic professional charts
        chart_sheet = wb.create_sheet("Charts")
        self._add_charts(wb, chart_sheet, df, query_language)

        # Save workbook
        wb.save(filepath)
        logger.info(f"Excel export completed: {filepath}")

        return str(filepath)

    def _detect_query_language(self, user_query: str) -> str:
        """Detect the primary language of the user query."""
        if not user_query:
            return "english"

        query_lower = user_query.lower()

        # Cyrillic characters indicate Russian
        cyrillic_count = sum(1 for char in query_lower if '\u0400' <= char <= '\u04ff')

        # Common Russian banking terms
        russian_terms = [
            'клиент', 'счет', 'баланс', 'операция', 'филиал', 'банк',
            'покажи', 'найди', 'все', 'сумма', 'общий'
        ]

        # Common Uzbek banking terms
        uzbek_terms = [
            'mijoz', 'hisob', 'balans', 'operatsiya', 'filial', 'bank',
            'ko\'rsat', 'toping', 'barchasi', 'miqdor', 'jami', 'so\'m'
        ]

        # Count term matches
        russian_matches = sum(1 for term in russian_terms if term in query_lower)
        uzbek_matches = sum(1 for term in uzbek_terms if term in query_lower)

        # Determine language
        if cyrillic_count > 0 or russian_matches > uzbek_matches:
            return "russian"
        elif uzbek_matches > 0:
            return "uzbek"
        else:
            return "english"

    def _format_currency_columns(self, df: pd.DataFrame, language: str = "english") -> pd.DataFrame:
        """Format currency columns from tiyin to proper currency display format."""

        # Common currency column patterns
        currency_patterns = [
            'amount', 'balance', 'limit', 'fee', 'salary', 'deposit',
            'withdrawal', 'payment', 'transfer', 'income', 'minimum',
            'maximum', 'overdraft', 'daily', 'monthly', 'annual'
        ]

        # Default currency display for zero amounts
        zero_display = "0.00 UZS" if language == "english" else (
            "0.00 сум" if language == "russian" else "0.00 so'm"
        )

        df_copy = df.copy()

        for col in df.columns:
            col_lower = str(col).lower()
            # Check if column contains currency amounts (likely in tiyin)
            if any(pattern in col_lower for pattern in currency_patterns):
                # Check if column contains numeric values that could be tiyin amounts
                if df[col].dtype in ['int64', 'float64'] and not df[col].isna().all():
                    # Convert tiyin to proper currency format for display
                    try:
                        df_copy[col] = df[col].apply(
                            lambda x: format_currency(int(x), language) if pd.notna(x) and x != 0 else zero_display
                        )
                    except (ValueError, TypeError):
                        # If conversion fails, keep original values
                        pass

        return df_copy

    def _write_data_sheet(self, ws, df: pd.DataFrame, query_info: Dict, language: str = "english"):
        """Write data to the main data sheet with professional formatting."""

        # Language-specific titles
        titles = {
            "english": {
                "title": "Bank Data Analysis Report"
            },
            "russian": {
                "title": "Отчёт по банковским данным"
            },
            "uzbek": {
                "title": "Bank ma'lumotlari tahlil hisoboti"
            }
        }

        lang_titles = titles.get(language, titles["english"])

        # Add professional report header
        ws.merge_cells('A1:E1')
        ws['A1'] = lang_titles["title"]
        ws['A1'].font = Font(size=16, bold=True, color=self.colors['primary'])
        ws['A1'].alignment = Alignment(horizontal='center')

        # Add data starting from row 3
        start_row = 3

        # Write headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=column)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx % 2 == 0:  # Alternate row coloring
                    cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def _create_summary_sheet(self, ws, df: pd.DataFrame, query_info: Dict, language: str = "english"):
        """Create a summary sheet with key statistics."""

        # Language-specific summary labels
        summary_labels = {
            "english": {
                "title": "Data Summary",
                "average": "Average",
                "total": "Total",
                "unique_values": "Unique Values"
            },
            "russian": {
                "title": "Сводка данных",
                "average": "Среднее",
                "total": "Всего",
                "unique_values": "Уникальные значения"
            },
            "uzbek": {
                "title": "Ma'lumotlar xulosasi",
                "average": "O'rtacha",
                "total": "Jami",
                "unique_values": "Noyob qiymatlar"
            }
        }

        labels = summary_labels.get(language, summary_labels["english"])

        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = labels["title"]
        ws['A1'].font = Font(size=14, bold=True, color=self.colors['primary'])
        ws['A1'].alignment = Alignment(horizontal='center')

        # Column statistics
        row = 2
        for col in df.columns:
            row += 1
            if df[col].dtype in ['int64', 'float64']:
                # Numeric column statistics (for non-currency columns)
                ws[f'A{row}'] = f"{col} ({labels['average']}):"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = round(df[col].mean(), 2) if not df[col].isna().all() else 0

                row += 1
                ws[f'A{row}'] = f"{col} ({labels['total']}):"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = round(df[col].sum(), 2) if not df[col].isna().all() else 0
            else:
                # Categorical column statistics (includes formatted currency columns)
                unique_count = df[col].nunique()
                ws[f'A{row}'] = f"{col} ({labels['unique_values']}):"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = unique_count

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _add_charts(self, wb, ws, df: pd.DataFrame, language: str = "english"):
        """Smart banking chart detection and dashboard creation."""

        # Language-specific sheet titles
        chart_titles = {
            "english": "Banking Dashboard",
            "russian": "Банковская панель",
            "uzbek": "Banking paneli"
        }

        ws.title = chart_titles.get(language, "Banking Dashboard")

        # Add professional dashboard header
        ws.merge_cells('A1:E1')
        ws['A1'] = chart_titles.get(language, "Banking Dashboard")
        ws['A1'].font = Font(size=16, bold=True, color=self.colors['primary'])
        ws['A1'].alignment = Alignment(horizontal='center')

        chart_row = 3

        try:
            # Smart banking data detection
            chart_insights = self._detect_banking_patterns(df)

            # Create dashboard based on detected patterns
            for insight in chart_insights:
                if insight['type'] == 'regional_analysis':
                    self._create_regional_chart(wb, ws, insight['data'], chart_row, language)
                    chart_row += 18
                elif insight['type'] == 'temporal_trends':
                    self._create_trend_chart(wb, ws, insight['data'], chart_row, language)
                    chart_row += 18
                elif insight['type'] == 'transaction_volumes':
                    self._create_volume_chart(wb, ws, insight['data'], chart_row, language)
                    chart_row += 18
                elif insight['type'] == 'risk_analysis':
                    self._create_risk_chart(wb, ws, insight['data'], chart_row, language)
                    chart_row += 18
                elif insight['type'] == 'financial_performance':
                    self._create_performance_chart(wb, ws, insight['data'], chart_row, language)
                    chart_row += 18

        except Exception as e:
            logger.warning(f"Error creating banking dashboard: {e}")
            # Add a note about chart creation failure
            ws['A3'] = f"Dashboard creation encountered an issue: {str(e)}"
            ws['A3'].font = Font(color='FF0000')

    def _detect_banking_patterns(self, df: pd.DataFrame) -> List[Dict]:
        """Smart chart selection based on data types and banking domain."""
        patterns = []

        if len(df) == 0:
            return patterns

        # Get column types with improved detection
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        categorical_cols = df.select_dtypes(include=['object', 'string']).columns.tolist()
        date_cols = [col for col in df.columns if any(term in col.lower() for term in ['date', 'time', 'created', 'month', 'year'])]

        # Enhanced numeric detection for calculated fields and currency columns
        for col in df.columns:
            col_lower = col.lower()
            # Check for currency/balance columns that might have been formatted as strings
            if any(term in col_lower for term in ['balance', 'amount', 'sum', 'total', 'price', 'cost', 'value', 'som']):
                if col not in numeric_cols and col not in categorical_cols:
                    # Try to detect if this could be treated as numeric data
                    sample_values = df[col].dropna().head(10)
                    if len(sample_values) > 0:
                        # Check if values look like formatted numbers or currency
                        if any(str(val).replace(',', '').replace('.', '').replace(' ', '').replace('UZS', '').replace('сум', '').replace('so\'m', '').isdigit() for val in sample_values):
                            categorical_cols.append(col)  # Treat as categorical for counting, but note it's numeric-ish

        # Chart Selection Logic - Try multiple chart types:

        # 1. LINE CHART: Date + Numeric (trends over time)
        if date_cols and numeric_cols:
            date_col = date_cols[0]
            numeric_col = numeric_cols[0]
            # Sort by date for proper line chart
            try:
                sorted_data = df[[date_col, numeric_col]].groupby(date_col)[numeric_col].sum().reset_index()
                if len(sorted_data) > 1:  # Need multiple points for line chart
                    patterns.append({
                        'type': 'temporal_trends',
                        'data': sorted_data,
                        'title': f"{numeric_col.replace('_', ' ').title()} Over Time"
                    })
            except Exception:
                pass  # Skip if aggregation fails

        # 2. BAR CHART: Categorical + Numeric (comparisons)
        if categorical_cols and numeric_cols:
            cat_col = categorical_cols[0]
            numeric_col = numeric_cols[0]
            try:
                # Aggregate data for bar chart
                bar_data = df.groupby(cat_col)[numeric_col].sum().reset_index()
                # Sort by value for better visualization
                bar_data = bar_data.sort_values(numeric_col, ascending=False).head(10)

                if len(bar_data) > 0:
                    # Determine chart type based on column names
                    if any(term in cat_col.lower() for term in ['region', 'branch', 'city']):
                        chart_type = 'regional_analysis'
                    elif any(term in cat_col.lower() for term in ['risk', 'rating', 'score']):
                        chart_type = 'risk_analysis'
                    elif any(term in cat_col.lower() for term in ['account', 'product', 'performance']):
                        chart_type = 'financial_performance'
                    else:
                        chart_type = 'regional_analysis'  # Default to bar chart

                    patterns.append({
                        'type': chart_type,
                        'data': bar_data,
                        'title': f"{numeric_col.replace('_', ' ').title()} by {cat_col.replace('_', ' ').title()}"
                    })
            except Exception:
                pass  # Skip if aggregation fails

        # 3. PIE CHART: Single categorical with counts (distributions)
        if len(categorical_cols) >= 1:
            cat_col = categorical_cols[0]
            try:
                # Get category counts for pie chart
                unique_values = df[cat_col].nunique()
                if 2 <= unique_values <= 10:  # Good range for pie chart
                    pie_data = df[cat_col].value_counts().head(8).reset_index()
                    pie_data.columns = [cat_col, 'count']
                    if len(pie_data) > 1:
                        patterns.append({
                            'type': 'transaction_volumes',
                            'data': pie_data,
                            'title': f"{cat_col.replace('_', ' ').title()} Distribution"
                        })
            except Exception:
                pass  # Skip if counting fails

        # GUARANTEED FALLBACK: Always create at least one chart if we have data
        if not patterns and len(df) > 0:
            try:
                # Emergency fallback: create a simple chart from any available data
                if len(df.columns) >= 2:
                    col1, col2 = df.columns[0], df.columns[1]

                    # If second column is numeric, create bar chart
                    if pd.api.types.is_numeric_dtype(df[col2]):
                        fallback_data = df[[col1, col2]].copy()
                        patterns.append({
                            'type': 'regional_analysis',
                            'data': fallback_data,
                            'title': f'{col2.replace("_", " ").title()} by {col1.replace("_", " ").title()}'
                        })
                    else:
                        # Create count chart of first column
                        count_data = df[col1].value_counts().reset_index()
                        count_data.columns = [col1, 'count']
                        patterns.append({
                            'type': 'transaction_volumes',
                            'data': count_data,
                            'title': f'Count by {col1.replace("_", " ").title()}'
                        })
                elif len(df.columns) == 1:
                    # Single column - create count chart
                    col = df.columns[0]
                    count_data = df[col].value_counts().reset_index()
                    count_data.columns = [col, 'count']
                    patterns.append({
                        'type': 'transaction_volumes',
                        'data': count_data,
                        'title': f'Count by {col.replace("_", " ").title()}'
                    })
                else:
                    # Ultimate fallback: create data overview chart
                    overview_data = pd.DataFrame({
                        'metric': ['Total Rows', 'Total Columns'],
                        'value': [len(df), len(df.columns)]
                    })
                    patterns.append({
                        'type': 'regional_analysis',
                        'data': overview_data,
                        'title': 'Data Overview'
                    })
            except Exception:
                # If even fallback fails, create minimal data summary
                try:
                    summary_data = pd.DataFrame({
                        'info': ['Rows', 'Columns'],
                        'count': [len(df), len(df.columns)]
                    })
                    patterns.append({
                        'type': 'regional_analysis',
                        'data': summary_data,
                        'title': 'Query Results Summary'
                    })
                except Exception:
                    pass  # Absolute last resort - no chart

        return patterns

    def _prepare_chart_data(self, df: pd.DataFrame, numeric_cols: List[str], categorical_cols: List[str]) -> pd.DataFrame:
        """Prepare aggregated data for charting."""

        if len(categorical_cols) > 0 and len(numeric_cols) > 0:
            # Group by first categorical column and aggregate numeric columns
            cat_col = categorical_cols[0]
            num_col = numeric_cols[0]

            if cat_col in df.columns and num_col in df.columns:
                grouped = df.groupby(cat_col)[num_col].agg(['sum', 'count', 'mean']).reset_index()
                grouped.columns = [cat_col, f'{num_col}_sum', f'{num_col}_count', f'{num_col}_mean']
                return grouped

        return df.head(20)  # Return first 20 rows if aggregation fails

    def _create_bar_chart(self, wb, ws, df: pd.DataFrame, category_col: str, value_col: str, start_row: int, language: str = "english"):
        """Create a professional bar chart."""

        # Language-specific labels
        chart_labels = {
            "english": {"chart_data": "Bar Chart Data"},
            "russian": {"chart_data": "Данные столбчатой диаграммы"},
            "uzbek": {"chart_data": "Ustunli diagramma ma'lumotlari"}
        }

        labels = chart_labels.get(language, chart_labels["english"])

        # Write chart data to worksheet
        ws.cell(start_row, 1, labels["chart_data"])
        ws.cell(start_row, 1).font = Font(bold=True)

        # Write headers
        ws.cell(start_row + 2, 1, category_col)
        ws.cell(start_row + 2, 2, value_col)

        # Write data (limit to top 10 for readability)
        chart_data = df.head(10)
        for idx, (_, row) in enumerate(chart_data.iterrows()):
            ws.cell(start_row + 3 + idx, 1, str(row[category_col])[:20])  # Truncate long names
            ws.cell(start_row + 3 + idx, 2, float(row[value_col]) if pd.notna(row[value_col]) else 0)

        # Use previously defined labels and add chart-specific terms
        labels["by"] = {"english": "by", "russian": "по", "uzbek": "bo'yicha"}[language]

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"{value_col} {labels['by']} {category_col}"
        chart.y_axis.title = value_col
        chart.x_axis.title = category_col

        # Set data range
        data_range = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 2 + len(chart_data))
        categories = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + 2 + len(chart_data))

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, f"D{start_row}")

    def _create_pie_chart(self, wb, ws, df: pd.DataFrame, category_col: str, start_row: int, language: str = "english"):
        """Create a professional pie chart."""

        # Count occurrences of each category
        if category_col in df.columns:
            value_counts = df[category_col].value_counts().head(8)  # Top 8 categories
        else:
            return

        # Write chart data
        ws.cell(start_row, 1, "Pie Chart Data")
        ws.cell(start_row, 1).font = Font(bold=True)

        ws.cell(start_row + 2, 1, category_col)
        ws.cell(start_row + 2, 2, "Count")

        for idx, (category, count) in enumerate(value_counts.items()):
            ws.cell(start_row + 3 + idx, 1, str(category)[:20])
            ws.cell(start_row + 3 + idx, 2, int(count))

        # Create chart
        chart = PieChart()
        chart.title = f"Distribution of {category_col}"

        # Set data range
        data_range = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 2 + len(value_counts))
        categories = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + 2 + len(value_counts))

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, f"D{start_row}")

    def _create_line_chart(self, wb, ws, df: pd.DataFrame, date_col: str, value_col: str, start_row: int, language: str = "english"):
        """Create a professional line chart for time series data."""

        # Sort by date and prepare data
        df_sorted = df.sort_values(date_col).head(20)

        # Write chart data
        ws.cell(start_row, 1, "Line Chart Data")
        ws.cell(start_row, 1).font = Font(bold=True)

        ws.cell(start_row + 2, 1, date_col)
        ws.cell(start_row + 2, 2, value_col)

        for idx, (_, row) in enumerate(df_sorted.iterrows()):
            ws.cell(start_row + 3 + idx, 1, str(row[date_col]))
            ws.cell(start_row + 3 + idx, 2, float(row[value_col]) if pd.notna(row[value_col]) else 0)

        # Create chart
        chart = LineChart()
        chart.title = f"{value_col} over {date_col}"
        chart.style = 12
        chart.y_axis.title = value_col
        chart.x_axis.title = date_col

        # Set data range
        data_range = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 2 + len(df_sorted))
        categories = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + 2 + len(df_sorted))

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, f"D{start_row}")

    def _has_date_column(self, df: pd.DataFrame) -> bool:
        """Check if dataframe has date/datetime columns."""
        for col in df.columns:
            if 'date' in col.lower() or 'time' in col.lower():
                return True
        return False

    def _get_date_column(self, df: pd.DataFrame) -> Optional[str]:
        """Get the first date column found."""
        for col in df.columns:
            if 'date' in col.lower() or 'time' in col.lower():
                return col
        return None

    def _create_regional_chart(self, wb, ws, chart_data: pd.DataFrame, start_row: int, language: str = "english"):
        """Create regional analysis chart with banking-specific styling."""
        if len(chart_data) == 0:
            return

        # Write chart data
        ws.cell(start_row, 1, "Regional Analysis")
        ws.cell(start_row, 1).font = Font(bold=True, color=self.colors['primary'])

        headers = list(chart_data.columns)
        for idx, header in enumerate(headers):
            ws.cell(start_row + 2, idx + 1, header)
            ws.cell(start_row + 2, idx + 1).font = Font(bold=True)

        for row_idx, (_, row) in enumerate(chart_data.iterrows()):
            for col_idx, value in enumerate(row):
                ws.cell(start_row + 3 + row_idx, col_idx + 1, value)

        # Create bar chart with banking colors
        chart = BarChart()
        chart.title = "Regional Distribution"
        chart.style = 10
        chart.height = 8
        chart.width = 12

        data_range = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 2 + len(chart_data))
        categories = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + 2 + len(chart_data))

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, f"D{start_row}")

    def _create_trend_chart(self, wb, ws, chart_data: pd.DataFrame, start_row: int, language: str = "english"):
        """Create temporal trend chart for banking analytics."""
        if len(chart_data) == 0:
            return

        ws.cell(start_row, 1, "Temporal Trends")
        ws.cell(start_row, 1).font = Font(bold=True, color=self.colors['primary'])

        # Write data headers
        headers = list(chart_data.columns)
        for idx, header in enumerate(headers):
            ws.cell(start_row + 2, idx + 1, header)
            ws.cell(start_row + 2, idx + 1).font = Font(bold=True)

        # Write data rows
        for row_idx, (_, row) in enumerate(chart_data.iterrows()):
            for col_idx, value in enumerate(row):
                ws.cell(start_row + 3 + row_idx, col_idx + 1, value)

        # Create line chart for trends
        chart = LineChart()
        chart.title = "Trends Over Time"
        chart.style = 12
        chart.height = 8
        chart.width = 12

        # Set data ranges
        data_range = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 2 + len(chart_data))
        categories = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + 2 + len(chart_data))

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, f"D{start_row}")

    def _create_volume_chart(self, wb, ws, chart_data: pd.DataFrame, start_row: int, language: str = "english"):
        """Create transaction volume chart."""
        if len(chart_data) == 0:
            return

        ws.cell(start_row, 1, "Transaction Volumes")
        ws.cell(start_row, 1).font = Font(bold=True, color=self.colors['primary'])

        # Write data headers
        headers = list(chart_data.columns)
        for idx, header in enumerate(headers):
            ws.cell(start_row + 2, idx + 1, header)
            ws.cell(start_row + 2, idx + 1).font = Font(bold=True)

        # Write data rows
        for row_idx, (_, row) in enumerate(chart_data.iterrows()):
            for col_idx, value in enumerate(row):
                ws.cell(start_row + 3 + row_idx, col_idx + 1, value)

        # Create pie chart for volume distribution
        chart = PieChart()
        chart.title = "Volume Distribution"
        chart.height = 8
        chart.width = 12

        # Set data ranges
        data_range = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 2 + len(chart_data))
        categories = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + 2 + len(chart_data))

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, f"D{start_row}")

    def _create_risk_chart(self, wb, ws, chart_data: pd.DataFrame, start_row: int, language: str = "english"):
        """Create risk analysis chart with professional styling."""
        if len(chart_data) == 0:
            return

        ws.cell(start_row, 1, "Risk Analysis")
        ws.cell(start_row, 1).font = Font(bold=True, color='FF0000')  # Red for risk

        # Write data headers
        headers = list(chart_data.columns)
        for idx, header in enumerate(headers):
            ws.cell(start_row + 2, idx + 1, header)
            ws.cell(start_row + 2, idx + 1).font = Font(bold=True)

        # Write data rows
        for row_idx, (_, row) in enumerate(chart_data.iterrows()):
            for col_idx, value in enumerate(row):
                ws.cell(start_row + 3 + row_idx, col_idx + 1, value)

        # Create specialized risk chart
        chart = BarChart()
        chart.title = "Risk Distribution"
        chart.style = 13  # Red style for risk
        chart.height = 8
        chart.width = 12

        # Set data ranges
        data_range = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 2 + len(chart_data))
        categories = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + 2 + len(chart_data))

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, f"D{start_row}")

    def _create_performance_chart(self, wb, ws, chart_data: pd.DataFrame, start_row: int, language: str = "english"):
        """Create financial performance dashboard chart."""
        if len(chart_data) == 0:
            return

        ws.cell(start_row, 1, "Financial Performance")
        ws.cell(start_row, 1).font = Font(bold=True, color=self.colors['success'])

        # Write data headers
        headers = list(chart_data.columns)
        for idx, header in enumerate(headers):
            ws.cell(start_row + 2, idx + 1, header)
            ws.cell(start_row + 2, idx + 1).font = Font(bold=True)

        # Write data rows
        for row_idx, (_, row) in enumerate(chart_data.iterrows()):
            for col_idx, value in enumerate(row):
                ws.cell(start_row + 3 + row_idx, col_idx + 1, value)

        # Create performance bar chart
        chart = BarChart()
        chart.title = "Performance Metrics"
        chart.style = 11  # Green style for performance
        chart.height = 8
        chart.width = 12

        # Set data ranges
        data_range = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 2 + len(chart_data))
        categories = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + 2 + len(chart_data))

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, f"D{start_row}")

    def export_to_excel(self, results_df: pd.DataFrame, filename: str, query_description: str, sql_query: str) -> str:
        """Export DataFrame to Excel with multilingual support."""
        # Detect language from query description
        language = self._detect_language(query_description)
        labels = self.labels[language]

        # Create workbook
        wb = Workbook()

        # Data sheet
        ws_data = wb.active
        ws_data.title = labels['data']
        self._write_data_sheet(ws_data, results_df, {
            'query': query_description,
            'sql': sql_query,
            'timestamp': datetime.now()
        }, language)

        # Charts sheet - always create if there's data
        if len(results_df) > 0:
            ws_charts = wb.create_sheet(title=labels['charts'])
            self._add_charts(wb, ws_charts, results_df, language)


        # Save file
        filepath = self.output_dir / filename
        wb.save(filepath)

        return str(filepath)


# Global Excel exporter instance
excel_exporter = ExcelExporter()